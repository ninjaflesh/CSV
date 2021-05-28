import sqlite3
import bcrypt
import openpyxl
from time import gmtime, strftime

# Подключение к БД
with sqlite3.connect('server.db') as db:
    sql = db.cursor()

sql.execute("""CREATE TABLE IF NOT EXISTS users (
    user_id INTEGER PRIMARY KEY,
    login TEXT,
    password TEXT,
    time TEXT
)""")
db.commit()

sql.execute("""CREATE TABLE IF NOT EXISTS grp (
    grp_id INTEGER PRIMARY KEY,
    grp_name TEXT,
    grp_num INTEGER
)""")
db.commit()

sql.execute("""CREATE TABLE IF NOT EXISTS audience (
    audience_id INTEGER PRIMARY KEY,
    audience_name TEXT,
    audience_num INTEGER,
    audience_type NUMERIC
)""")
db.commit()

sql.execute("""CREATE TABLE IF NOT EXISTS work (
    work_id INTEGER PRIMARY KEY,
    work_name TEXT,
    work_type NUMERIC,
    work_p TEXT NOT NULL
)""")
db.commit()

# Подключение к книги Exel
wb = openpyxl.load_workbook('schedule.xlsx')


class User:  # Клас User, принемает логин и пароль
    def __init__(self, login: str, password: str):
        self.user_login = login
        self.user_password = password
        self.create_time = strftime("%Y-%m-%d %H:%M:%S", gmtime())

    # Метод шифрует пароль
    def get_hashed_password(self, password):
        return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())

    # Метод шифрует пароль и проверяет с ранее шифрованным
    def check_password(self, hash_p):
        return bcrypt.checkpw(self.user_password.encode('utf-8'), hash_p)

    # Метод проверяет логин на сущестование
    def check_login(self):
        sql.execute(
            f"SELECT login, password FROM users WHERE login = '{self.user_login}'")
        coll = sql.fetchone()
        if coll is not None:
            if self.check_password(coll[1]):
                print('Вы успешно авторизовались.')
                self.edit_schedule()
            else:
                print("Не верный логин или пароль.")
        else:
            print("Создайте новую запись.")
            self.register()  # Отправляем на регистрацию

    # Метод регистрирует пользователя
    def register(self):
        new_login = input('Введите логи для регистрации: ')

        sql.execute(
            f"SELECT login FROM users WHERE login = '{new_login}'")
        if sql.fetchone() is None:
            new_password = input('Введите пароль для регистрации: ')

            sql.execute(f"INSERT INTO users VALUES (?, ?, ?, ?)",
                        (None, new_login, self.get_hashed_password(new_password), self.create_time))
            db.commit()
            print('Запись создана.')
        else:
            print('Логин занят.')

    # Метод позволяющий пользователю взаимодействовать с расписанием.
    def edit_schedule(self):
        print("Для добавления группы введите: 1\nДля добавления аудитории введите: 2\n"
              "Для добавления предмета введите: 3\nДля удаления группы введите: 4\n"
              "Для удаления аудитории введите: 5\nДля удаления предмета введите: 6\n"
              "Для создания расписания введите: 7")
        check = int(input('Выберите действие: '))

        if check == 1:
            create_group()
        elif check == 2:
            create_audience()
        elif check == 3:
            create_work()
        elif check == 4:
            delet_group()
        elif check == 5:
            delet_audience()
        elif check == 6:
            delet_work()
        elif check == 7:
            build_schedule()
        else:
            print("Я тебя не понимаю О_о")


def create_group():
    group = input('Введите название группы: ')
    group_num = input('Введите количество человек в группе: ')

    sql.execute(f"INSERT INTO grp VALUES (?, ?, ?)",
                (None, group, group_num))
    db.commit()
    print(f"Добавлена группа {group}.")


def create_audience():
    audience = input('Введите номер аудитории: ')
    audience_num = input('Введите вместимость адитории: ')
    audience_type = input('Введите тип аудитории [0 - лаба, 1 - обычная]: ')

    sql.execute(f"INSERT INTO audience VALUES (?, ?, ?, ?)",
                (None, audience, audience_num, audience_type))
    db.commit()
    print(f"Добавлена аудитория №{audience}.")


def create_work():
    work = input('Введите название предмета: ')
    work_type = input(
        'Введите тип аудитории для проведения занятия [0 - лаба, 1 - обычная]: ')
    work_p = input('Введите имя преподавателя: ')

    sql.execute(f"INSERT INTO work VALUES (?, ?, ?, ?)",
                (None, work, work_type, work_p))
    db.commit()
    print(f"Добавлен предмет {work}.")


def delet_group():
    sql.execute(f"SELECT * FROM grp ")
    for x in sql.fetchall():
        print(x)
    id = int(input("Укажите id группы для удаления: "))
    sql.execute(f"DELETE FROM grp WHERE grp_id = '{id}'")
    db.commit()
    print("Группа удалена.")


def delet_audience():
    sql.execute(f"SELECT * FROM audience ")
    for x in sql.fetchall():
        print(x)
    id = int(input("Укажите id аудитории для удаления: "))
    sql.execute(f"DELETE FROM audience WHERE audience_id = '{id}'")
    db.commit()
    print("Аудитория удалена.")


def delet_work():
    sql.execute(f"SELECT * FROM work ")
    for x in sql.fetchall():
        print(x)
    id = int(input("Укажите id предмета для удаления: "))
    sql.execute(f"DELETE FROM work WHERE work_id = '{id}'")
    db.commit()
    print("Предмет удален.")


def build_schedule():
    sql.execute(
        f"SELECT grp_name, grp_num FROM grp")
    group = sql.fetchall()

    sql.execute(
        f"SELECT audience_name, audience_num, audience_type FROM audience")
    audience = sql.fetchall()

    sql.execute(
        f"SELECT work_name, work_type FROM work")
    work = sql.fetchall()

    group.sort(key=lambda group: group[1], reverse=True)
    audience = sorted(audience, key=lambda type: type[1], reverse=True)
    audience.sort(key=lambda type: type[2])
    work.sort(key=lambda type: type[1])

    data = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота']
    time = ['8:00 - 11:20', '11:30 - 15:00', '15:10 - 18:30', '18:45 -  чч:мм']

    sheets = wb.sheetnames
    sheet = wb[sheets[0]]

    s1 = 1
    for row in range(1, len(audience) + 1):
        cell = sheet.cell(row=s1, column=1)
        value = audience[row-1][0]
        cell.value = value

        for col in range(2, len(data) + 2):
            cell2 = sheet.cell(row=s1, column=col)
            value2 = data[col-2]
            cell2.value = value2

        for t in range(1, len(time) + 1):
            cell3 = sheet.cell(row=s1+t, column=1)
            value3 = time[t-1]
            cell3.value = value3

        s1 = s1 + 6
    wb.save('schedule.xlsx')

    def search(name_a):
        for rowr in sheet.iter_rows():  # Проходимся по строчкам таблицы
            for cell in rowr:  # Проходимся по ячейкам строчки
                if cell.value == name_a:  # Проверяем ячейку на наличие аудитории
                    return cell.row

    for x in group:  # Идем по группам
        s2 = 0
        cl = 1
        for j in data:  # Идем по дням
            p1 = -1
            p2 = -1
            p3 = -1
            p4 = -1
            for i in range(2):  # Лимит промежутков в день
                f = False
                c = -1
                while f == False:
                    c = c + 1
                    if c < len(audience):
                        if x[1] <= audience[c][1]:  # Проверяем влезет ли группа в аудиторию
                            r = search(audience[c][0])
                            if s2 < len(work):
                                if audience[c][2] == work[s2][1]:
                                    if sheet.cell(row=r+1, column=1+cl).value == None:
                                        p1 = p1 + 1
                                        if p1 < 1:
                                            cell4 = sheet.cell(
                                                row=r+1, column=1+cl)
                                            value4 = str(
                                                x[0]) + ' ' + str(work[s2][0])
                                            cell4.value = value4
                                            s2 = s2 + 1
                                            f = True
                                    elif sheet.cell(row=r+2, column=1+cl).value == None:
                                        p2 = p2 + 1
                                        if p2 < 1:
                                            cell4 = sheet.cell(
                                                row=r+2, column=1+cl)
                                            value4 = str(
                                                x[0]) + ' ' + str(work[s2][0])
                                            cell4.value = value4
                                            s2 = s2 + 1
                                            f = True
                                    elif sheet.cell(row=r+3, column=1+cl).value == None:
                                        p3 = p3 + 1
                                        if p3 < 1:
                                            cell4 = sheet.cell(
                                                row=r+3, column=1+cl)
                                            value4 = str(
                                                x[0]) + ' ' + str(work[s2][0])
                                            cell4.value = value4
                                            s2 = s2 + 1
                                            f = True
                                    elif sheet.cell(row=r+4, column=1+cl).value == None:
                                        p4 = p4 + 1
                                        if p4 < 1:
                                            cell4 = sheet.cell(
                                                row=r+4, column=1+cl)
                                            value4 = str(
                                                x[0]) + ' ' + str(work[s2][0])
                                            cell4.value = value4
                                            s2 = s2 + 1
                                            f = True
                    else:
                        f = True
            cl = cl + 1
        work.append(work[0])
        work.remove(work[0])
        wb.save('schedule.xlsx')


t1 = User('Admin', "admin")
t1.check_login()
